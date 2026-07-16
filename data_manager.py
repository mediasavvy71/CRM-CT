"""
data_manager.py
----------------
Excel-backed data access layer for the Chronic Trace CRM.

Every screen in the app reads and writes through this module instead of
touching openpyxl/pandas directly. This keeps the UI layer (app.py) simple
and makes the data logic testable without a display.

Workbook layout (ChronicTraceCRM.xlsx):
    Organizations   - accounts / prospects / customers
    Contacts        - people, linked to an Organization via OrgID
    Opportunities   - sales pipeline (schema reserved for a future phase)
    Activities      - calls/emails/meetings/demos (schema reserved)
    FollowUps       - reminders / due dates (schema reserved)

Only Organizations and Contacts have working CRUD in this phase. The other
sheets are created up front so later phases don't require a migration.
"""

from __future__ import annotations

import os
import shutil
import uuid
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

DATE_FMT = "%Y-%m-%d"
DATETIME_FMT = "%Y-%m-%d %H:%M"

ORG_TYPES = ["High School", "College", "Healthcare", "Enterprise", "Partnership", "Other"]

SHEET_COLUMNS = {
    "Organizations": [
        "OrgID", "Name", "Type", "Industry", "Address", "City", "State", "Zip",
        "Phone", "Website", "Notes", "CreatedDate", "LastContactDate",
    ],
    "Contacts": [
        "ContactID", "OrgID", "OrgName", "FirstName", "LastName", "Title",
        "Email", "Phone", "DecisionMaker", "Notes", "CreatedDate", "LastContactDate",
    ],
    # Reserved for future phases:
    "Opportunities": [
        "OppID", "OrgID", "OrgName", "Name", "Stage", "Amount", "Probability",
        "ExpectedCloseDate", "Source", "Notes", "CreatedDate",
    ],
    "Activities": [
        "ActivityID", "OrgID", "ContactID", "OppID", "Type", "Date",
        "Subject", "Notes",
    ],
    "FollowUps": [
        "FollowUpID", "OrgID", "ContactID", "DueDate", "Description",
        "Status", "CreatedDate",
    ],
}

HEADER_FILL = PatternFill(start_color="1F2A44", end_color="1F2A44", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _now_str() -> str:
    return datetime.now().strftime(DATE_FMT)


def new_id(prefix: str) -> str:
    return f"{prefix}-{uuid.uuid4().hex[:8].upper()}"


class CRMWorkbook:
    """Thin wrapper around an Excel workbook used as the CRM's datastore."""

    def __init__(self, path: str):
        self.path = os.path.expanduser(path)
        if not os.path.exists(self.path):
            self._create_new_workbook()

    # ------------------------------------------------------------------ #
    # Workbook bootstrap
    # ------------------------------------------------------------------ #
    def _create_new_workbook(self) -> None:
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)  # drop default "Sheet"

        for sheet_name, columns in SHEET_COLUMNS.items():
            ws = wb.create_sheet(sheet_name)
            ws.append(columns)
            for col_idx, _ in enumerate(columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="left")
                ws.column_dimensions[get_column_letter(col_idx)].width = 20

        wb.save(self.path)

    def backup(self) -> str:
        """Copy the current workbook to a timestamped backup before writing."""
        backup_dir = os.path.join(os.path.dirname(self.path), "backups")
        os.makedirs(backup_dir, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = os.path.splitext(os.path.basename(self.path))[0]
        backup_path = os.path.join(backup_dir, f"{base}_{stamp}.xlsx")
        shutil.copy2(self.path, backup_path)
        return backup_path

    # ------------------------------------------------------------------ #
    # Generic sheet IO
    # ------------------------------------------------------------------ #
    def read_sheet(self, sheet_name: str) -> pd.DataFrame:
        columns = SHEET_COLUMNS[sheet_name]
        try:
            df = pd.read_excel(self.path, sheet_name=sheet_name, dtype=str)
        except ValueError:
            # Sheet missing (older workbook) -- create it on the fly.
            self._ensure_sheet(sheet_name)
            df = pd.read_excel(self.path, sheet_name=sheet_name, dtype=str)
        for col in columns:
            if col not in df.columns:
                df[col] = ""
        df = df[columns].fillna("")
        return df

    def _ensure_sheet(self, sheet_name: str) -> None:
        wb = load_workbook(self.path)
        if sheet_name in wb.sheetnames:
            return
        ws = wb.create_sheet(sheet_name)
        ws.append(SHEET_COLUMNS[sheet_name])
        wb.save(self.path)

    def _write_sheet(self, sheet_name: str, df: pd.DataFrame) -> None:
        """Overwrite a single sheet while preserving the other sheets."""
        wb = load_workbook(self.path)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        columns = SHEET_COLUMNS[sheet_name]
        ws.append(columns)
        for col_idx, _ in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

        for _, row in df.iterrows():
            ws.append([row.get(c, "") for c in columns])

        wb.save(self.path)

    # ------------------------------------------------------------------ #
    # Organizations
    # ------------------------------------------------------------------ #
    def list_organizations(self) -> pd.DataFrame:
        return self.read_sheet("Organizations")

    def get_organization(self, org_id: str) -> Optional[dict]:
        df = self.list_organizations()
        match = df[df["OrgID"] == org_id]
        return None if match.empty else match.iloc[0].to_dict()

    def add_organization(self, data: dict) -> str:
        df = self.list_organizations()
        org_id = new_id("ORG")
        row = {c: data.get(c, "") for c in SHEET_COLUMNS["Organizations"]}
        row["OrgID"] = org_id
        row["CreatedDate"] = _now_str()
        row.setdefault("LastContactDate", "")
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
        self._write_sheet("Organizations", df)
        return org_id

    def update_organization(self, org_id: str, data: dict) -> None:
        df = self.list_organizations()
        idx = df.index[df["OrgID"] == org_id]
        if idx.empty:
            raise KeyError(f"Organization {org_id} not found")
        for k, v in data.items():
            if k in df.columns and k not in ("OrgID", "CreatedDate"):
                df.loc[idx, k] = v
        self._write_sheet("Organizations", df)
        # keep denormalized OrgName in sync on Contacts
        if "Name" in data:
            self._sync_org_name(org_id, data["Name"])

    def delete_organization(self, org_id: str) -> None:
        df = self.list_organizations()
        df = df[df["OrgID"] != org_id]
        self._write_sheet("Organizations", df)
        # Unlink contacts rather than deleting them
        contacts = self.list_contacts()
        mask = contacts["OrgID"] == org_id
        contacts.loc[mask, "OrgID"] = ""
        contacts.loc[mask, "OrgName"] = ""
        self._write_sheet("Contacts", contacts)

    def _sync_org_name(self, org_id: str, new_name: str) -> None:
        contacts = self.list_contacts()
        mask = contacts["OrgID"] == org_id
        if mask.any():
            contacts.loc[mask, "OrgName"] = new_name
            self._write_sheet("Contacts", contacts)

    # ------------------------------------------------------------------ #
    # Contacts
    # ------------------------------------------------------------------ #
    def list_contacts(self) -> pd.DataFrame:
        return self.read_sheet("Contacts")

    def get_contact(self, contact_id: str) -> Optional[dict]:
        df = self.list_contacts()
        match = df[df["ContactID"] == contact_id]
        return None if match.empty else match.iloc[0].to_dict()

    def contacts_for_org(self, org_id: str) -> pd.DataFrame:
        df = self.list_contacts()
        return df[df["OrgID"] == org_id]

    def add_contact(self, data: dict) -> str:
        df = self.list_contacts()
        contact_id = new_id("CON")
        row = {c: data.get(c, "") for c in SHEET_COLUMNS["Contacts"]}
        row["ContactID"] = contact_id
        row["CreatedDate"] = _now_str()
        row.setdefault("LastContactDate", "")
        if row.get("OrgID"):
            org = self.get_organization(row["OrgID"])
            row["OrgName"] = org["Name"] if org else ""
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
        self._write_sheet("Contacts", df)
        return contact_id

    def update_contact(self, contact_id: str, data: dict) -> None:
        df = self.list_contacts()
        idx = df.index[df["ContactID"] == contact_id]
        if idx.empty:
            raise KeyError(f"Contact {contact_id} not found")
        for k, v in data.items():
            if k in df.columns and k not in ("ContactID", "CreatedDate"):
                df.loc[idx, k] = v
        if "OrgID" in data:
            org = self.get_organization(data["OrgID"]) if data["OrgID"] else None
            df.loc[idx, "OrgName"] = org["Name"] if org else ""
        self._write_sheet("Contacts", df)

    def delete_contact(self, contact_id: str) -> None:
        df = self.list_contacts()
        df = df[df["ContactID"] != contact_id]
        self._write_sheet("Contacts", df)

    # ------------------------------------------------------------------ #
    # Dashboard metrics
    # ------------------------------------------------------------------ #
    def dashboard_metrics(self) -> dict:
        orgs = self.list_organizations()
        contacts = self.list_contacts()

        by_type = orgs["Type"].replace("", "Unspecified").value_counts().to_dict()
        decision_makers = contacts[contacts["DecisionMaker"].str.lower().isin(["y", "yes", "true"])]

        recent_orgs = orgs.sort_values("CreatedDate", ascending=False).head(5)
        recent_contacts = contacts.sort_values("CreatedDate", ascending=False).head(5)

        return {
            "total_orgs": len(orgs),
            "total_contacts": len(contacts),
            "decision_makers": len(decision_makers),
            "orgs_by_type": by_type,
            "recent_orgs": recent_orgs.to_dict("records"),
            "recent_contacts": recent_contacts.to_dict("records"),
        }
